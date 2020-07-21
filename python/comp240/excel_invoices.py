from openpyxl import load_workbook, Workbook

def dictionaryBuilder():
    db = {}
    wb = load_workbook('database_month_2.xlsx')
    for ws in wb.sheetnames:
        db[ws] = {}
        for row in wb[ws].iter_rows(min_row=3, min_col=1):
            if row[0].value is not None:
                col = 0
                db[ws][row[0].value] = {}
                for cell in row:
                    db[ws][row[0].value][wb[ws][2][col].value] = cell.value
                    col += 1
    return db

# print(dictionaryBuilder())


def clientInvoices(invoice_no):
    db = dictionaryBuilder()
    total = db['invoices'][invoice_no]['total_price']
    products = []
    # customer name
    customer = db['invoices'][invoice_no]
    customerid = customer['customer_id']
    name = f"{db['customers'][customerid]['f_name']} {db['customers'][customerid]['l_name']}"

    # items purchased
    for items in db['invoice line items']:
        if db['invoice line items'][items]['invoice_no'] == invoice_no:
            individualProducts =db['invoice line items'][items]['product_id']
            products.append(f"{db['product'][individualProducts]['name']} x {db['invoice line items'][items]['quantity']}---  ${db['invoice line items'][items]['price']}")

    # return name, total, products
    return {'Invoice': customer,
            'Name': name,
            'Purchased': products,
            'total spent': total}


