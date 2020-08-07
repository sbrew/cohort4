from openpyxl import load_workbook, Workbook


def salesTracker():
    database = {}
    customertTable = {}
    invoiceTable = {}
    invoiceLineTable = {}
    productTable = {}
    sheets = []
    wb = load_workbook('database_month_2.xlsx')

    for tabs in wb.sheetnames:
        sheets.append(tabs)

    for row in wb['customers'].iter_rows(min_row=2, max_col=wb['customers'].max_column, max_row=wb['customers'].max_row, values_only=True):
        customer_id = row[0]
        data = {'customer_id': row[0],
                'f_name': row[1],
                'l_name': row[2],
                'phone': row[3],
                'email': row[4],
                'street': row[5],
                'city': row[6],
                'zipcode': row[7]
                }
        customertTable[customer_id] = data
        database[sheets[0]] = customertTable

    for row in wb['invoices'].iter_rows(min_row=2, max_col=wb['invoices'].max_column, max_row=wb['invoices'].max_row, values_only=True):
        invoice_no = row[0]
        data = {'invoice_no': row[0],
                "customer_id": (row[1]),
                "invoice_date": (row[2]),
                "payment_method": row[3],
                "total_price": row[4]
                }
        invoiceTable[invoice_no] = data
        database[sheets[1]] = invoiceTable

    for row in wb['invoice line items'].iter_rows(min_row=2, max_col=wb['invoice line items'].max_column, max_row=wb['invoice line items'].max_row, values_only=True):
        invoice_line_id = row[0]
        data = {'invoice_no': row[1],
                'product_id': (row[2]),
                'quantity': (row[3]),
                'price': row[4]
                }
        invoiceLineTable[invoice_line_id] = data
        database[sheets[2]] = invoiceLineTable

    for row in wb['product'].iter_rows(min_row=2, max_col=wb['product'].max_column, max_row=wb['product'].max_row, values_only=True):
        product_id = row[0]
        data = {'product_id': row[0],
                'name': (row[1]),
                'description': (row[2]),
                'quantity': row[3],
                'cost': row[4]
                }
        productTable[product_id] = data
        database[sheets[3]] = productTable

    return database


def dictionaryBuilder():
    db = {}
    wb = load_workbook('database_month_2.xlsx')
    for ws in wb.sheetnames:
        db[ws] = {}
        for row in wb[ws].iter_rows(min_row=3, min_col=1):
            if row[0].value is not None:
                col = 0
                db[ws][row[0].value] = {}
                for cell in row:
                    db[ws][row[0].value][wb[ws][2][col].value] = cell.value
                    col += 1
    return db

# print(dictionaryBuilder())


def clientInvoices(invoice_no):
    db = dictionaryBuilder()
    name = ""
    # print(db)
    # print(invoice_no)
    total = db['invoices'][invoice_no]['total_price']
    products = []
    # customer name
    customer = db['invoices'][invoice_no]
    customerid = customer['customer_id']
    name = f"{db['customers'][customerid]['f_name']} {db['customers'][customerid]['l_name']}"
    # invoicenumber
    invoiceNumber = db['invoices'][invoice_no]['invoice_no']

    # items purchased
    for items in db['invoice line items']:
        if db['invoice line items'][items]['invoice_no'] == invoice_no:
            individualProducts =db['invoice line items'][items]['product_id']
            products.append(f"{db['product'][individualProducts]['name']} x {db['invoice line items'][items]['quantity']}---  ${db['invoice line items'][items]['price']}")

    with open('report.txt', 'w') as report:
        report.write("--------Customer invoice--------\n")
        report.write(f"Customer: {name}       Date:{db['invoices'][invoice_no]['invoice_date']}\n")
        report.write(f"Customer email: {db['customers'][customerid]['email']}\n")
        report.write(f"Invoice number: {invoiceNumber}\n")
        report.write("\n----------------------\n")
        report.write("Items:\n")
        for items in products:
            report.write(f"{items}\n")
        report.write(f"total: ${total}\n")
        report.write("\n----------------------\n")
        
    return name, total, products
x=int(input("Please enter an invoice number: "))
# invoice = input()
clientInvoices(x)